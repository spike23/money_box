import os
from openpyxl import load_workbook
from openpyxl.styles import Side, Border, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


class ExcelWriter:

    COLUMNS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

    def __init__(self, db_tool_obj):
        self.db_tool_obj = db_tool_obj

    @property
    def path(self):
        path: str = os.path.dirname(__file__).replace('excel', 'template')
        return path

    @property
    def template_file_path(self) -> str:
        file_path: str = os.path.join(self.path, 'money_calendar_template.xlsx')
        return file_path

    @property
    def file_path(self):
        file_path: str = os.path.join(self.path, 'money_calendar.xlsx')
        return file_path

    @property
    def paid_days(self):
        return self.db_tool_obj.get_all_done()

    def __set_background_color(self, sheet: Worksheet, paid_days: list) -> None:
        """
        set_background_color for xlsx cells
        :param sheet: xlsx sheet
        :param paid_days: all days those were paid
        :return: None
        """
        for row in range(1, 38):
            cols = self.COLUMNS if row != 37 else self.COLUMNS[:5]
            for idx in cols:
                color = 'DAF7A6' if sheet[f'{idx}{row}'].value in paid_days else 'FCF89F'
                side = Side(border_style='thin', color='000000')
                border = Border(top=side, right=side, left=side, bottom=side)
                sheet[f'{idx}{row}'].fill = PatternFill(bgColor=color, fill_type='solid', start_color=color,
                                                        end_color=color, fgColor=color)
                sheet[f'{idx}{row}'].border = border

    def write(self) -> None:
        """
        write xlsx file
        :return: None
        """
        wb = load_workbook(self.template_file_path)
        sheet = wb['year_days']
        self.__set_background_color(sheet, self.paid_days)
        wb.save(self.file_path)
        wb.close()
